import os
import re
import io
import base64
import time
import datetime
import traceback
import pandas as pd
import pymysql
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from fpdf import FPDF
import openpyxl

# External Data
from Courses import KEYWORDS, SKILLS_DICT, COURSES, JOB_DESCRIPTIONS
from celery import Celery

# Internal Utils
from utils import pdf_reader, clean_text_nltk, extract_resume_data, parse_resume_sections, predict_field_fast, calculate_rigorous_score, mask_pii

# Import the background task
from batch_selector import process_batch_task

# --- CONFIGURATION ---
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024   # 5 MB hard limit
app.secret_key = 'super_secret_key'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- CELERY CONFIGURATION ---
def make_celery(app):
    celery = Celery(
        app.import_name,
        backend=app.config['CELERY_RESULT_BACKEND'],
        broker=app.config['CELERY_BROKER_URL']
    )
    celery.conf.update(app.config)
    class ContextTask(celery.Task):
        def __call__(self, *args, **kwargs):
            with app.app_context():
                return self.run(*args, **kwargs)
    celery.Task = ContextTask
    return celery

app.config.update(
    CELERY_BROKER_URL='redis://localhost:6379/0',
    CELERY_RESULT_BACKEND='redis://localhost:6379/0'
)

celery = make_celery(app)

# --- DATABASE CONNECTION ---
try:
    connection = pymysql.connect(host='localhost', user='root', password='')
    cursor = connection.cursor()
except Exception as e:
    print(f"DB Connection Error: {e}")
    connection = None
    cursor = None

if connection:
    cursor.execute("CREATE DATABASE IF NOT EXISTS cv")
    cursor.execute("USE cv")
    
    table_sql = """
    CREATE TABLE IF NOT EXISTS user_data (
        ID INT NOT NULL AUTO_INCREMENT,
        Name varchar(500) NOT NULL,
        Email_ID VARCHAR(500) NOT NULL,
        resume_score VARCHAR(8) NOT NULL,
        Timestamp VARCHAR(50) NOT NULL,
        Page_no VARCHAR(5) NOT NULL,
        Predicted_Field BLOB NOT NULL,
        User_level BLOB NOT NULL,
        Recommended_skills BLOB NOT NULL,
        Recommended_courses BLOB NOT NULL,
        PRIMARY KEY (ID)
    );
    """
    cursor.execute(table_sql)

    skills_table_sql = """
    CREATE TABLE IF NOT EXISTS user_skills (
        ID INT NOT NULL AUTO_INCREMENT,
        user_id INT NOT NULL,
        skill_name VARCHAR(100) NOT NULL,
        PRIMARY KEY (ID),
        FOREIGN KEY (user_id) REFERENCES user_data(ID) ON DELETE CASCADE
    );
    """
    cursor.execute(skills_table_sql)

# --- HELPER FUNCTIONS ---

def cleanup_file(filepath):
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            print(f"Cleaned up file: {filepath}")
    except Exception as e:
        print(f"Error deleting file {filepath}: {e}")

def get_image_file(img_data):
    if not img_data or not isinstance(img_data, str): return None
    try:
        if img_data.startswith('data:image'):
            header, encoded = img_data.split(",", 1)
            data = base64.b64decode(encoded)
            return io.BytesIO(data)
        elif os.path.exists(img_data):
            return img_data
    except Exception as e:
        print(f"Error processing image: {e}")
        return None
    return None

# --- PDF GENERATION FUNCTIONS ---
def create_classic_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    def sanitize(text):
        if not text: return ""
        text = str(text)
        text = re.sub(r'[^\x00-\x7F]', ' ', text)
        text = text.replace('\r', '').replace('\n', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def get_data(key): return sanitize(data.get(key.lower()) or data.get(key, ''))
    
    name = get_data('name'); email = get_data('email'); mobile = get_data('mobile')
    photo = data.get('photo'); summary = get_data('summary'); experience = get_data('experience')
    education = get_data('education'); projects = get_data('projects'); skills = get_data('skills')

    if photo:
        img_file = get_image_file(photo)
        if img_file: pdf.image(img_file, x=140, y=15, w=45, h=45)

    pdf.set_font("Times", 'B', 24); pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 15, name if name else "YOUR NAME", 0, 1, 'L')
    pdf.set_font("Times", '', 12); pdf.set_text_color(80, 80, 80)
    contact_line = ""
    if email: contact_line += email
    if mobile: contact_line += (" | " if contact_line else "") + mobile
    pdf.cell(0, 6, contact_line, 0, 1, 'L')
    pdf.set_draw_color(0, 0, 0)
    pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2); pdf.ln(10)

    def add_section_header(title):
        pdf.ln(5); pdf.set_font("Times", 'B', 14); pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, title.upper(), 0, 1, 'L')
        pdf.set_draw_color(0, 0, 0); pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.set_draw_color(255, 255, 255); pdf.ln(5)

    def add_section_text(text):
        pdf.set_font("Times", '', 11); pdf.multi_cell(0, 6, text)

    if summary and summary != "[No Summary details found]":
        add_section_header("Objective/Summary"); add_section_text(summary)
    if skills and skills != "[No Skills details found]":
        add_section_header("Skills"); add_section_text(skills)
    if experience and experience != "[No Experience details found]":
        add_section_header("Work Experience")
        lines = experience.split('.'); pdf.set_font("Times", 'B', 12)
        for line in lines:
            if not line.strip(): continue
            if len(line) < 70: pdf.multi_cell(0, 6, line); pdf.ln(2)
            else: pdf.set_font("Times", '', 11); pdf.multi_cell(0, 6, line); pdf.ln(2)
            pdf.set_font("Times", 'B', 12)
    if projects and projects != "[No Projects details found]":
        add_section_header("Projects"); add_section_text(projects)
    if education and education != "[No Education details found]":
        add_section_header("Education"); add_section_text(education)
    return pdf

def create_modern_pdf(data):
    pdf = FPDF()
    pdf.add_page()
    BRAND_BLUE = (41, 98, 255); BRAND_GRAY = (100, 100, 100)
    pdf.set_auto_page_break(auto=True, margin=15)

    def sanitize(text):
        if not text: return ""
        text = str(text)
        text = re.sub(r'[^\x00-\x7F]', ' ', text)
        text = text.replace('\r', '').replace('\n', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def get_data(key): return sanitize(data.get(key.lower()) or data.get(key, ''))

    name = get_data('name'); email = get_data('email'); mobile = get_data('mobile')
    photo = data.get('photo'); summary = get_data('summary'); experience = get_data('experience')
    education = get_data('education'); projects = get_data('projects'); skills = get_data('skills')

    if photo:
        img_file = get_image_file(photo)
        if img_file: pdf.image(img_file, x=150, y=15, w=40, h=40)

    pdf.set_font("Arial", 'B', 28); pdf.set_text_color(*BRAND_BLUE)
    pdf.cell(0, 15, name if name else "YOUR NAME", 0, 1, 'L')
    pdf.set_font("Arial", '', 11); pdf.set_text_color(*BRAND_GRAY)
    contact_line = ""
    if email: contact_line += email
    if mobile: contact_line += (" • " if contact_line else "") + mobile
    pdf.cell(0, 6, contact_line, 0, 1, 'L')
    pdf.set_draw_color(*BRAND_BLUE); pdf.set_fill_color(*BRAND_BLUE)
    pdf.rect(10, pdf.get_y() + 2, 190, 2, 'F'); pdf.ln(10)

    def add_section_header(title):
        pdf.ln(6); pdf.set_font("Arial", 'B', 16); pdf.set_text_color(*BRAND_BLUE)
        pdf.cell(0, 8, title, 0, 1, 'L'); pdf.ln(2)

    def add_section_text(text):
        pdf.set_font("Arial", '', 11); pdf.set_text_color(50, 50, 50); pdf.multi_cell(0, 6, text)

    if summary and summary != "[No Summary details found]":
        add_section_header("About Me"); add_section_text(summary)
    if skills and skills != "[No Skills details found]":
        add_section_header("Technical Skills"); add_section_text(skills)
    if experience and experience != "[No Experience details found]":
        add_section_header("Professional Experience")
        lines = experience.split('.'); pdf.set_font("Arial", 'B', 12); pdf.set_text_color(30, 30, 30)
        for line in lines:
            if not line.strip(): continue
            if len(line) < 70: pdf.multi_cell(0, 6, line); pdf.ln(2)
            else: pdf.set_font("Arial", '', 11); pdf.set_text_color(50, 50, 50); pdf.multi_cell(0, 6, line); pdf.ln(2)
            pdf.set_font("Arial", 'B', 12)
    if projects and projects != "[No Projects details found]":
        add_section_header("Key Projects"); add_section_text(projects)
    if education and education != "[No Education details found]":
        add_section_header("Education"); add_section_text(education)
    return pdf

# --- ROUTES ---
@app.route('/')
def index(): return render_template('home.html')

@app.route('/user')
def user_view(): return render_template('user.html')

@app.route('/admin')
def admin_login_view(): return render_template('admin_login.html')

@app.route('/dashboard')
def admin_dashboard_view(): return render_template('admin_dashboard.html')

@app.route('/generator')
def generator_view(): return render_template('generator.html')

@app.route('/shortlist')
def shortlist_view(): return render_template('shortlist.html')

@app.route('/api/analytics/skills_gap')
def analytics_skills_gap():
    if not connection: return jsonify([])
    cursor.execute("SELECT skill_name, COUNT(*) as count FROM user_skills GROUP BY skill_name ORDER BY count DESC LIMIT 10")
    results = cursor.fetchall()
    return jsonify([{'skill': row[0], 'count': row[1]} for row in results])

@app.route('/api/analytics/high_potential')
def analytics_high_potential():
    if not connection: return jsonify([])
    cursor.execute("SELECT Name, Email_ID, resume_score, Predicted_Field FROM user_data WHERE (Page_no = '1' OR User_level = 'Fresher') AND CAST(resume_score AS DECIMAL) > 80 ORDER BY CAST(resume_score AS DECIMAL) DESC")
    results = cursor.fetchall()
    return jsonify([{'name': r[0], 'email': r[1], 'score': r[2], 'field': r[3]} for r in results])

@app.route('/api/start_batch', methods=['POST'])
def start_batch_process():
    BATCH_FOLDER = 'batch_resumes'
    if not os.path.exists(BATCH_FOLDER):
        return jsonify({'error': 'Batch folder "batch_resumes" does not exist.'}), 404
    task = process_batch_task.apply_async()
    return jsonify({'task_id': task.id, 'state': task.state, 'status': 'Batch processing started in background.'})

@app.route('/api/batch_status/<task_id>')
def batch_status(task_id):
    task = process_batch_task.AsyncResult(task_id)
    if task.state == 'PENDING':
        response = {'state': task.state, 'status': 'Pending...'}
    elif task.state != 'FAILURE':
        response = {'state': task.state, 'current': task.info.get('current', 0), 'total': task.info.get('total', 1), 'status': task.info.get('status', '')}
        if task.state == 'SUCCESS': response['result'] = task.info
    else:
        response = {'state': task.state, 'status': str(task.info)}
    return jsonify(response)

@app.route('/api/download_batch_report')
def download_batch_report():
    OUTPUT_FILE = 'outputs/batch_report.xlsx'
    if not os.path.exists(OUTPUT_FILE):
        return jsonify({'error': 'Report not generated yet.'}), 404
    return send_file(OUTPUT_FILE, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Candidate_Rankings.xlsx')

@app.route('/api/shortlist', methods=['POST'])
def shortlist_resumes():
    jd = request.form.get('job_description', '')
    files = request.files.getlist('resumes')
    results = []
    target_jd = jd if jd else JOB_DESCRIPTIONS.get('default', '')

    for file in files:
        if file.filename == '': continue
        filepath = None
        try:
            safe_name = f"{int(time.time()*1000)}_{secure_filename(file.filename)}"
            filepath  = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
            file.save(filepath)
            resume_text_raw, pages = pdf_reader(filepath)
            cleanup_file(filepath); filepath = None
            if len(resume_text_raw.strip()) < 50:
                results.append({'name': file.filename, 'email': 'N/A', 'score': 0, 'status': 'Error: No Text'}); continue
            resume_text_clean = clean_text_nltk(resume_text_raw)
            data = extract_resume_data(resume_text_raw, resume_text_clean)
            sections = parse_resume_sections(resume_text_raw)
            score_data = calculate_rigorous_score(
                resume_text=resume_text_raw, resume_edu_section=sections.get('Education', ''),
                resume_exp_section=sections.get('Experience', ''), user_skills=data.get('skills', []),
                job_description_text=target_jd)
            results.append({'filename': file.filename, 'name': data.get('name', 'Unknown'),
                'email': data.get('email', 'N/A'), 'score': score_data['total_score'],
                'breakdown': score_data['breakdown'], 'skills': data.get('skills', [])})
        except Exception as e:
            print(f"Error processing {file.filename}: {e}"); traceback.print_exc()
            results.append({'name': file.filename, 'email': 'Error', 'score': 0, 'status': 'Processing Failed'})
        finally:
            if filepath: cleanup_file(filepath)

    results.sort(key=lambda x: x['score'], reverse=True)
    return jsonify(results)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    filepath = None
    try:
        # ── 1. Validate file presence ──────────────────────────────────
        if 'file' not in request.files:
            return jsonify({'error': 'No file part in the request.'}), 400
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'No file selected.'}), 400

        # ── 2. Validate file type ──────────────────────────────────────
        ext = os.path.splitext(file.filename)[1].lower()
        if ext != '.pdf' and file.mimetype not in ('application/pdf', 'application/x-pdf'):
            return jsonify({'error': 'Only PDF files are accepted.'}), 415

        # ── 3. Save to disk with a safe unique name ────────────────────
        #       BUG 1 FIX IS HERE ↓
        safe_name = f"{int(time.time()*1000)}_{secure_filename(file.filename)}"
        filepath  = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        file.save(filepath)

        # ── 4. Parse PDF — pass path string, NOT FileStorage object ───
        try:
            resume_text_raw, pages = pdf_reader(filepath)   # ← path string ✓
        except Exception as pdf_err:
            return jsonify({'error': f'PDF parsing failed: {str(pdf_err)}'}), 422

        if len(resume_text_raw.strip()) < 50:
            return jsonify({'error': 'Not enough text in PDF. It may be image-only or empty.'}), 422

        # ── 5. NLP processing ──────────────────────────────────────────
        resume_text_clean = clean_text_nltk(resume_text_raw)
        data              = extract_resume_data(resume_text_raw, resume_text_clean)

        name       = data.get('name',          'Unknown')
        email      = data.get('email',         'Unknown')
        mobile     = data.get('mobile_number', 'Unknown')
        skills     = data.get('skills',        [])
        cand_level = 'Fresher' if pages == 1 else ('Intermediate' if pages == 2 else 'Experienced')
        sections   = parse_resume_sections(resume_text_raw)

        # ── 6. Field prediction & recommendations ──────────────────────
        reco_field         = predict_field_fast(resume_text_clean)
        recommended_skills = SKILLS_DICT.get(reco_field, [])
        rec_courses        = COURSES.get(reco_field, [])

        # ── 7. Score calculation ───────────────────────────────────────
        job_description_input = request.form.get('job_description', '')
        target_jd = job_description_input if job_description_input else JOB_DESCRIPTIONS.get('default', '')

        score_data = calculate_rigorous_score(
            resume_text=resume_text_raw,
            resume_edu_section=sections.get('Education', ''),
            resume_exp_section=sections.get('Experience', ''),
            user_skills=skills,
            job_description_text=target_jd
        )
        score = score_data['total_score']

                # ── 8. Database insert — UPDATED TO SAVE BREAKDOWN ──────
        if connection:
            try:
                ts        = time.time()
                timestamp = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d_%H:%M:%S')
                
                # Extract breakdown values safely
                s_skill = str(score_data['breakdown'].get('skills', 0))
                s_edu   = str(score_data['breakdown'].get('education', 0))
                s_exp   = str(score_data['breakdown'].get('experience', 0))
                s_sem   = str(score_data['breakdown'].get('semantic_fit', 0))

                insert_sql = """INSERT INTO user_data
                    (Name, Email_ID, resume_score, Timestamp, Page_no,
                     Predicted_Field, User_level, Recommended_skills, Recommended_courses,
                     Skills_Score, Edu_Score, Exp_Score, Semantic_Score)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
                    
                cursor.execute(insert_sql, (
                    name, email, str(score), timestamp, str(pages),
                    reco_field, cand_level, str(recommended_skills), str(rec_courses),
                    s_skill, s_edu, s_exp, s_sem
                ))
                
                user_id = cursor.lastrowid
                if skills:
                    cursor.executemany(
                        "INSERT INTO user_skills (user_id, skill_name) VALUES (%s, %s)",
                        [(user_id, skill) for skill in skills]
                    )
                connection.commit()
            except Exception as db_err:
                print(f"[DB WARNING] Insert failed (non-fatal): {db_err}")
                try: connection.rollback()
                except Exception: pass

        # ── 9. Return result ───────────────────────────────────────────
        return jsonify({
            'name': name, 'email': email, 'mobile': mobile,
            'pages': pages, 'level': cand_level, 'skills': skills,
            'field': reco_field, 'rec_skills': recommended_skills,
            'rec_courses': rec_courses, 'score': score,
            'breakdown': score_data['breakdown']
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': f'Server error: {str(e)}'}), 500

    finally:
        if filepath:
            cleanup_file(filepath)   # always delete temp file


@app.route('/api/login', methods=['POST'])
def admin_login():
    d = request.json
    if d.get('username') == 'dipen' and d.get('password') == 'dipen123': return jsonify({'success': True})
    return jsonify({'success': False}), 401

@app.route('/api/data')
def admin_data():
    if not connection: return jsonify([])
    privacy_mode = request.args.get('privacy', 'false').lower() == 'true'
    cursor.execute("SELECT * FROM user_data")
    raw_data = cursor.fetchall()
    formatted_data = []
    for row in raw_data:
        new_row = []
        for i, item in enumerate(row):
            new_row.append(item.decode('utf-8') if isinstance(item, bytes) else item)
        if privacy_mode:
            new_row[1], new_row[2] = mask_pii(new_row[1], new_row[2])
        formatted_data.append(new_row)
    return jsonify(formatted_data)

@app.route('/api/download')
def download_file():
    if not connection: return "No DB connection"
    cursor.execute("SELECT * FROM user_data")
    raw_data = cursor.fetchall()
    formatted_data = [[item.decode('utf-8') if isinstance(item, bytes) else item for item in row] for row in raw_data]
    columns = [desc[0] for desc in cursor.description]
    df = pd.DataFrame(formatted_data, columns=columns)
    output = io.StringIO(); df.to_csv(output, index=False); output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv', as_attachment=True, download_name='user_data.csv')

@app.route('/api/parse_for_generator', methods=['POST'])
def parse_for_generator():
    if 'file' not in request.files: return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    filepath = None
    try:
        safe_name = f"{int(time.time()*1000)}_{secure_filename(file.filename)}"
        filepath  = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        file.save(filepath)
        resume_text_raw, pages = pdf_reader(filepath)
        if not resume_text_raw or len(resume_text_raw.strip()) < 5:
            return jsonify({'error': 'Could not extract text', 'name': 'Unknown', 'email': '', 'mobile': '', 'summary': '', 'experience': '', 'education': '', 'projects': '', 'skills': ''}), 200
        resume_text_clean = clean_text_nltk(resume_text_raw)
        basic_data = extract_resume_data(resume_text_raw, resume_text_clean)
        sections = parse_resume_sections(resume_text_raw)
        return jsonify({
            'name': basic_data.get('name'), 'email': basic_data.get('email'),
            'mobile': basic_data.get('mobile_number'),
            'summary': sections.get('Summary', ''), 'experience': sections.get('Experience', ''),
            'education': sections.get('Education', ''), 'projects': sections.get('Projects', ''),
            'skills': "\n".join(basic_data.get('skills', []))
        })
    except Exception as e:
        return jsonify({'error': str(e), 'name': 'Unknown', 'email': '', 'mobile': '', 'summary': '', 'experience': '', 'education': '', 'projects': '', 'skills': ''}), 200
    finally:
        if filepath: cleanup_file(filepath)

@app.route('/api/analyze_updated_resume', methods=['POST'])
def analyze_updated_resume():
    data = request.json
    reconstructed_text = f"{data.get('summary','')} {data.get('experience','')} {data.get('education','')} {data.get('projects','')} {data.get('skills','')}"
    cleaned_text = clean_text_nltk(reconstructed_text)
    reco_field = predict_field_fast(cleaned_text)
    recommended_skills = SKILLS_DICT.get(reco_field, [])
    score_categories = {'Career Objective': ['objective','summary'], 'Declaration': ['declaration'], 'Hobbies': ['hobbies'], 'Achievements': ['achievements'], 'Projects': ['projects']}
    score = sum(20 for kws in score_categories.values() if any(k in cleaned_text for k in kws))
    return jsonify({'score': score, 'field': reco_field, 'rec_skills': recommended_skills, 'status': 'Analyzed Successfully'})

@app.route('/api/llm/optimize', methods=['POST'])
def llm_optimize():
    data = request.json
    text_to_optimize = data.get('text', '')
    if not text_to_optimize: return jsonify({'optimized': ''})
    optimized = text_to_optimize
    if "responsible for" in text_to_optimize.lower():
        optimized = text_to_optimize.replace("Responsible for", "Managed") + " (Improved by System)"
    return jsonify({'optimized': optimized})

@app.route('/api/download_updated_resume', methods=['POST'])
def download_updated_resume():
    data = request.json
    template_style = data.get('style', 'classic')
    try:
        pdf = create_modern_pdf(data) if template_style == 'modern' else create_classic_pdf(data)
        try: pdf_str = pdf.output(dest='S').encode('latin-1')
        except UnicodeEncodeError: pdf_str = pdf.output(dest='S').encode('utf-8')
        file_stream = io.BytesIO(pdf_str); file_stream.seek(0)
        filename = data.get('name', 'Resume').replace(" ", "_") + f"_{template_style}.pdf"
        return send_file(file_stream, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': 'Failed to generate PDF', 'details': str(e)}), 500

@app.route('/api/candidate/<int:id>', methods=['DELETE'])
def delete_candidate(id):
    if not connection: return jsonify({'error': 'Database connection error'}), 500
    try:
        cursor.execute("DELETE FROM user_data WHERE ID = %s", (id,))
        connection.commit()
        return jsonify({'success': True, 'message': 'Candidate deleted successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.errorhandler(413)
def too_large(e):
    return jsonify({'error': 'File too large. Maximum allowed size is 5 MB.'}), 413

if __name__ == '__main__':
    app.run(debug=True)